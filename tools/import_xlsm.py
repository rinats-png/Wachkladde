#!/usr/bin/env python3
"""
Migration: Wachkladde_<Monat>.xlsm  ->  wachkladde.json

    pip install openpyxl
    python3 tools/import_xlsm.py Wachkladde_08_August_2026.xlsm > wachkladde.json

Gelesen werden:
  * EingabeSchicht  -> Dienstgruppen, Beamte, Amtsbezeichnungen, Auswahllisten, Rotation
  * Tagesblätter    -> je Blatt zwei Schichten (TD Spalten A..K, ND Spalten M..W)

Die Zieldatei wird in der HTML-Anwendung unter
"Einstellungen -> Daten -> JSON importieren" eingelesen.
"""
import json, sys, uuid, datetime
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

BLOCKS = [("TD", 0), ("ND", 12)]          # Spaltenversatz des Nachtdienst-Blocks
BESTAND_ROWS = list(range(45, 56))
LISTEN_RANGES = {"abwesenheit": ("L", 53, 74), "einteilung": ("L", 78, 101),
                 "einsatzzug": ("Q", 53, 55)}
ABST_VORLAGEN = [
    {"titel": "Wachbesetzung / Streife",
     "felder": ["Anlass", "Beamte", "L- Wache", "Wache", "Streife"]},
    {"titel": "Abstellung / BSOD",
     "felder": ["Anlass", "Bea.", "Name", "MOZ", "Dienstanz.", "bes. FEM"]},
]


def uid(p):
    return f"{p}_{uuid.uuid4().hex[:7]}"


def first_day(wb):
    for ws in wb.worksheets:
        if ws.title.strip().isdigit():
            return ws.title
    return wb.sheetnames[1]


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)     # berechnete Werte
    es = wb["EingabeSchicht"]
    tpl = wb[first_day(wb)]

    # ---- Stammdaten: 5 Dienstgruppen, Namensspalten B/E/H/K/N, Zeilen 10..25
    dgs = []
    for i, c in enumerate(["B", "E", "H", "K", "N"], start=1):
        amt = get_column_letter(column_index_from_string(c) + 1)
        beamte = []
        for r in range(10, 26):
            name = es[f"{c}{r}"].value
            if not name or not str(name).strip():
                continue
            beamte.append({"id": uid("b"), "name": str(name).strip(),
                           "amtsbez": str(es[f"{amt}{r}"].value or "NIT").strip(),
                           "rolle": "praktikant" if r == 25 else "beamter", "aktiv": True})
        dgs.append({"id": uid("dg"), "nr": i, "name": f"DG {i}", "beamte": beamte})

    # ---- Auswahllisten
    listen = {}
    for key, (c, r0, r1) in LISTEN_RANGES.items():
        vals = [str(es[f"{c}{r}"].value).strip() for r in range(r0, r1 + 1)
                if es[f"{c}{r}"].value not in (None, "")]
        listen[key] = [""] + [v for v in vals if v]
    r0 = next((r for r in range(1, tpl.max_row + 1)
               if "Verwarngeld D219" in str(tpl[f"A{r}"].value or "")), 45)
    r1 = next((r for r in range(r0, tpl.max_row + 1)
               if "übergebender DGL" in str(tpl[f"A{r}"].value or "")), r0 + 11)
    listen["bestand"] = [str(tpl[f"A{r}"].value or "").strip() for r in range(r0, r1)]

    # ---- Rotation: Kette über DG-Nummern (nicht über Wochentage!)
    def rot(ck, cv, r0=47, r1=51):
        m = {}
        for r in range(r0, r1 + 1):
            k, v = es[f"{ck}{r}"].value, es[f"{cv}{r}"].value
            if isinstance(k, (int, float)) and isinstance(v, (int, float)):
                m[str(int(k))] = int(v)
        return m

    start = es["H2"].value
    start = start.date() if isinstance(start, datetime.datetime) else datetime.date.today()
    rotation = {"startDatum": start.isoformat(),
                "startDgTag": int(tpl["K1"].value or 1),
                "nextTag": rot("Q", "R"), "nachtVonTag": rot("K", "L"),
                "verstVonTag": rot("Z", "AA"), "verstNachtVonVerst": rot("W", "X")}

    # ---- Tagesblätter (layout-robust über Ankertexte statt fester Zeilen)
    tage = {}
    for ws in wb.worksheets:
        head = ws.title.strip().split()[0]
        if not head.isdigit():
            continue
        try:
            datum = start.replace(day=1) + datetime.timedelta(days=int(head) - 1)
        except ValueError:
            continue
        key = datum.isoformat()
        tage[key] = {}
        for schicht, off in BLOCKS:
            def cell(letter, row, off=off):
                return ws[f"{get_column_letter(column_index_from_string(letter) + off)}{row}"].value

            def txt(letter, row, off=off):
                v = cell(letter, row, off)
                return "" if v is None else str(v).strip()

            hits = {k: [] for k in ("erg", "lehr", "sonst", "best", "dgl")}
            for r in range(1, ws.max_row + 1):
                a = txt("A", r).lower()
                for name, needle in (("erg", "ergänzungsdienst"), ("lehr", "lehrgänge"),
                                     ("sonst", "sonstiges"), ("best", "verwarngeld d219"),
                                     ("dgl", "übergebender dgl")):
                    if needle in a:
                        hits[name].append(r)
            anchors = {k: v[0] for k, v in hits.items() if v}
            # "übergebender DGL" steht in manchen Blättern auch VOR der Bestandsliste
            if "best" in anchors:
                nach = [r for r in hits["dgl"] if r > anchors["best"]]
                anchors["dgl"] = nach[0] if nach else ws.max_row + 1

            r_pers = next((r for r in range(1, ws.max_row + 1)
                           if cell("A", r) == 1), 6)
            r_erg = anchors.get("erg", 22)
            r_lehr = anchors.get("lehr", 30)
            r_sonst = anchors.get("sonst", 37)
            r_best = anchors.get("best", 45)
            r_dgl = anchors.get("dgl", 56)

            dgnr = cell("K", 1)
            dg = next((d for d in dgs if d["nr"] == dgnr), None)
            aktive = [b for b in (dg["beamte"] if dg else []) if b["aktiv"]]
            eint = {}
            for i, b in enumerate(aktive):
                r = r_pers + i
                if r >= r_erg:
                    break
                eint[b["id"]] = {"abwesenheit": txt("E", r), "einteilung": txt("G", r),
                                 "bemerkung": txt("H", r)}
            erg = [{"name": txt("B", r), "amtsbez": "", "abwesenheit": txt("E", r),
                    "einteilung": txt("G", r), "bemerkung": txt("H", r)}
                   for r in range(r_erg + 1, r_lehr) if txt("B", r)]

            # Abstellungen: Blöcke werden an "Anlass:" in Spalte I erkannt
            abst, cur = [], None
            for r in range(1, r_best):
                lab = txt("I", r).rstrip(":").strip()
                if not lab:
                    continue
                if lab.lower() == "anlass":
                    cur = {"titel": "", "felder": [], "werte": {}}
                    abst.append(cur)
                if cur is not None:
                    cur["felder"].append(lab)
                    cur["werte"][lab] = txt("J", r)
            for a in abst:
                a["titel"] = ("Wachbesetzung / Streife" if "L- Wache" in a["felder"]
                              else "Abstellung / BSOD")

            lehr = [{"name": txt("A", r), "art": txt("C", r), "ort": txt("G", r),
                     "zeit": txt("J", r), "eintrag": txt("K", r)}
                    for r in range(r_lehr + 2, r_sonst)]
            sonst = "\n".join(t for t in (txt("A", r) for r in range(r_sonst, r_best))
                              if t and not t.lower().startswith("sonstiges"))
            bestand = {}
            for r in range(r_best, r_dgl):
                label = txt("A", r, 0)
                if label:
                    bestand[label] = txt("H", r)
            ergdg = cell("A", r_erg + 1)
            tage[key][schicht] = {
                "datum": key, "schicht": schicht, "dgNr": dgnr,
                "einsatzzug": txt("H", 4), "einteilungen": eint,
                "ergaenzung": {"dgNr": ergdg if isinstance(ergdg, int) else None,
                               "eintraege": erg},
                "abstellungen": [{"vorlage": None, "titel": a["titel"], "felder": a["felder"],
                                  "werte": a["werte"]} for a in abst],
                "lehrgaenge": lehr, "sonstiges": sonst, "bestand": bestand,
                "dglAb": txt("H", r_dgl), "dglAn": txt("K", r_dgl)}

    out = {"schemaVersion": 1,
           "meta": {"revier": str(tpl["A1"].value or "Polizeirevier"),
                    "titel": "Wachdienstplan für"},
           "settings": {"user": "", "rolle": "dgl", "syncMode": "off", "syncUrl": "",
                        "syncInt": 10, "colors": {}, "datum": start.isoformat()},
           "stammdaten": {"dienstgruppen": dgs, "listen": listen,
                          "abstVorlagen": ABST_VORLAGEN},
           "rotation": rotation, "tage": tage, "_cursor": 0}
    json.dump(out, sys.stdout, ensure_ascii=False, indent=1)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Aufruf: python3 tools/import_xlsm.py <Datei.xlsm> > wachkladde.json")
    main(sys.argv[1])
